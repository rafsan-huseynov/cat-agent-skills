#!/usr/bin/env python3
"""Extract tables from a PDF into a formatted workbook (or CSV) using pdfplumber.

Usage:
    python extract_tables.py document.pdf [--out-dir out]
                                          [--pages 2-5]
                                          [--contains rebate]
                                          [--prefix contoso]
                                          [--format xlsx|csv|both]

By default this writes a single **.xlsx** workbook with one nicely-formatted
sheet (tab) per detected table — bold frozen header row, an autofilter, and
sensible column widths — which handles multiple tables and page splits cleanly.
Pass `--format csv` for one CSV per table instead, or `--format both` for both.
A JSON summary of everything produced (files, sheets, source page, row/column
counts) is printed to stdout so the calling agent can report back accurately.

This is a deterministic post-processing helper: it does not invent, reformat, or
"correct" cell values — it only lifts the table grid out of the PDF and cleans
away fully empty rows and stray whitespace (empty columns are preserved). Values
are written as text to preserve them verbatim; the formatting is purely visual.
"""
import argparse
import csv
import json
import os
import re
import sys


def parse_pages(spec, page_count):
    """Turn a '2', '2-5' or '1,3,5-7' spec into a sorted list of 0-based indexes."""
    if not spec:
        return list(range(page_count))
    wanted = set()
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            lo, hi = part.split("-", 1)
            for n in range(int(lo), int(hi) + 1):
                wanted.add(n - 1)
        else:
            wanted.add(int(part) - 1)
    return sorted(p for p in wanted if 0 <= p < page_count)


def clean_table(rows):
    """Trim whitespace and drop fully empty rows (empty columns are kept)."""
    grid = [[("" if c is None else str(c).strip()) for c in row] for row in rows]
    grid = [row for row in grid if any(cell for cell in row)]
    if not grid:
        return grid
    width = max(len(row) for row in grid)
    return [row + [""] * (width - len(row)) for row in grid]


def slugify(text, fallback):
    text = re.sub(r"[^a-z0-9]+", "-", (text or "").lower()).strip("-")
    return text or fallback


def sheet_title(base, page_number, table_number, used):
    """Build a unique, Excel-legal sheet name (<=31 chars, no []:*?/\\)."""
    raw = f"{base}-p{page_number}-t{table_number}"
    clean = re.sub(r"[\[\]:*?/\\]", "-", raw)[:31] or "Sheet"
    candidate, suffix = clean, 1
    while candidate.lower() in used:
        tag = f"-{suffix}"
        candidate = clean[: 31 - len(tag)] + tag
        suffix += 1
    used.add(candidate.lower())
    return candidate


def write_csv(path, grid):
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        csv.writer(fh).writerows(grid)


def style_sheet(worksheet, grid):
    """Apply a bold frozen header, autofilter, and column widths."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    columns = len(grid[0])
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(columns)}{len(grid)}"
    )
    for col in range(columns):
        longest = max((len(row[col]) for row in grid), default=0)
        worksheet.column_dimensions[get_column_letter(col + 1)].width = min(
            max(longest + 2, 10), 60
        )


def main(argv=None):
    parser = argparse.ArgumentParser(description="Extract PDF tables to xlsx/csv.")
    parser.add_argument("pdf", help="path to the source PDF")
    parser.add_argument("--out-dir", default="out", help="directory for output files")
    parser.add_argument("--pages", default="", help="page selector, e.g. 2-5 or 1,3")
    parser.add_argument("--contains", default="", help="keep only tables containing this text (case-insensitive)")
    parser.add_argument("--prefix", default="", help="filename prefix for the output")
    parser.add_argument(
        "--format",
        choices=["xlsx", "csv", "both"],
        default="xlsx",
        help="output format (default: xlsx, one formatted tab per table)",
    )
    args = parser.parse_args(argv)

    try:
        import pdfplumber  # noqa: WPS433 (imported lazily so the message is clear)
    except ImportError:
        sys.exit(
            "pdfplumber is required. Install it with `pip install pdfplumber`, "
            "or extract the tables manually following the skill's cleaning rules."
        )

    want_xlsx = args.format in ("xlsx", "both")
    workbook = None
    if want_xlsx:
        try:
            from openpyxl import Workbook
        except ImportError:
            if args.format == "xlsx":
                sys.stderr.write(
                    "openpyxl not available; falling back to CSV output. "
                    "Install it with `pip install openpyxl` for formatted .xlsx.\n"
                )
                args.format = "csv"
                want_xlsx = False
            else:
                sys.stderr.write(
                    "openpyxl not available; producing CSV only.\n"
                )
                want_xlsx = False
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)

    want_csv = args.format in ("csv", "both")

    if not os.path.isfile(args.pdf):
        sys.exit(f"file not found: {args.pdf}")

    os.makedirs(args.out_dir, exist_ok=True)
    base = args.prefix or slugify(os.path.splitext(os.path.basename(args.pdf))[0], "document")
    needle = args.contains.lower()

    tables_out, csv_files, used_titles = [], [], set()
    with pdfplumber.open(args.pdf) as pdf:
        page_indexes = parse_pages(args.pages, len(pdf.pages))
        for page_index in page_indexes:
            page = pdf.pages[page_index]
            page_number = page_index + 1
            tables = page.extract_tables() or []
            for table_number, raw in enumerate(tables, start=1):
                grid = clean_table(raw)
                if not grid:
                    continue
                if needle and needle not in "\n".join(
                    " ".join(row) for row in grid
                ).lower():
                    continue

                entry = {"page": page_number, "rows": len(grid), "columns": len(grid[0])}

                if want_csv:
                    csv_name = f"{base}-p{page_number}-t{table_number}.csv"
                    csv_path = os.path.join(args.out_dir, csv_name)
                    write_csv(csv_path, grid)
                    csv_files.append(csv_path)
                    entry["csv"] = csv_path

                if want_xlsx:
                    title = sheet_title(base, page_number, table_number, used_titles)
                    worksheet = workbook.create_sheet(title=title)
                    for row in grid:
                        worksheet.append(row)
                    style_sheet(worksheet, grid)
                    entry["sheet"] = title

                tables_out.append(entry)

    xlsx_path = None
    if want_xlsx and tables_out:
        xlsx_path = os.path.join(args.out_dir, f"{base}.xlsx")
        workbook.save(xlsx_path)

    summary = {
        "source": args.pdf,
        "count": len(tables_out),
        "xlsx": xlsx_path,
        "csv_files": csv_files,
        "tables": tables_out,
    }
    json.dump(summary, sys.stdout, indent=2)
    sys.stdout.write("\n")
    if not tables_out:
        sys.stderr.write(
            "No tables matched. Try a different --pages range, drop --contains, "
            "or verify the PDF has extractable (non-image) tables.\n"
        )
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
